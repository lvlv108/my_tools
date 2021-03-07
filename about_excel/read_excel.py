# -*- coding: utf-8 -*-
# @Time:2021/3/7 6:49 下午
# @Author:lvlv
# @File:read_excel.py
from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
for i in range(1,31):
    print(sheet_ranges.cell(columggtgtn=1, row=i).value)