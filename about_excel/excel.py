# -*- coding: utf-8 -*-
# @Time:2021/3/7 6:23 下午
# @Author:lvlv
# @File:excel.py.py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()  # 实例化workbook
dest_filename = 'empty_book.xlsx'
ws1 = wb.active  # 通过active调起正在运行的工作表
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title="my_sheet")
for i in range(1,31):
    ws4.cell(column= 1 ,row = i ,value = 'test')


wb.save(filename=dest_filename)
